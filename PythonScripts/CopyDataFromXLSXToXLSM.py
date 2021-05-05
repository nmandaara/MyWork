
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected
         

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

#! Python 3
# - Copy and Paste Ranges using OpenPyXl library

import openpyxl
import sys

n=len(sys.argv)
print("Total arguments passed:",n)

File=sys.argv[1]
Path=sys.argv[2]
NewFile=File[:-5]+".xlsm"
#Prepare the spreadsheets to copy from and paste too.


#File to be copied
wb = openpyxl.load_workbook(File) #Add file name
sheet = wb.get_sheet_by_name("Data") #Add Sheet name

#File to be pasted into
template = openpyxl.load_workbook("GIVE YOUR XLSM FILE ",keep_vba=True) #Add file name
temp_sheet = template.get_sheet_by_name("Data Sheet") #Add Sheet name

#def createData()
print("Processing...")
selectedRange = copyRange(1,1,255,10000,sheet) #Change the 4 number values
pastingRange = pasteRange(1,1,255,10000,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too
template.save(NewFile)
print("Range copied and pasted!")

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.

